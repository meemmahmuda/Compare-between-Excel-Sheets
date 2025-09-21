# import pandas as pd

# df1 = pd.read_excel("dbbl.xlsx", sheet_name="Sheet1")
# df2 = pd.read_excel("dncc.xlsx", sheet_name="Sheet1")


# merged = pd.merge(df1, df2, left_on='Txn ID', right_on='Tranno', how='outer', indicator=True)



# matched = merged[merged['_merge'] == 'both']
# unmatched = merged[merged['_merge'] != 'both']


# unmatched.to_excel("unmatched_both.xlsx", index=False)

# print("Done! Files created: matched_both.xlsx, unmatched_both.xlsx")


# import pandas as pd
# from tkinter import Tk, filedialog, simpledialog
# import os

# root = Tk()
# root.withdraw()

# def load_file(file_path):
#     ext = os.path.splitext(file_path)[1].lower()
#     if ext == ".csv":
#         return pd.read_csv(file_path)
#     elif ext == ".xlsx":
#         return pd.read_excel(file_path, sheet_name=0)
#     else:
#         raise ValueError(f"Unsupported file type: {ext}")

# file1 = filedialog.askopenfilename(
#     title="Select first file",
#     filetypes=[("Excel or CSV files", "*.xlsx *.csv")]
# )

# file2 = filedialog.askopenfilename(
#     title="Select second file",
#     filetypes=[("Excel or CSV files", "*.xlsx *.csv")]
# )

# df1 = load_file(file1)
# df2 = load_file(file2)

# df1_columns_lower = {col.lower(): col for col in df1.columns}
# df2_columns_lower = {col.lower(): col for col in df2.columns}

# col1_input = simpledialog.askstring("Input", "Enter column name from first file to compare:").strip().lower()
# col2_input = simpledialog.askstring("Input", "Enter column name from second file to compare:").strip().lower()

# col1 = df1_columns_lower.get(col1_input)
# col2 = df2_columns_lower.get(col2_input)

# if not col1 or not col2:
#     raise ValueError("Column not found in one of the files. Check spelling.")

# merged = pd.merge(df1, df2, left_on=col1, right_on=col2, how='outer', indicator=True)

# unmatched = merged[merged['_merge'] != 'both']

# unmatched_file = filedialog.asksaveasfilename(
#     title="Save unmatched file as",
#     defaultextension=".xlsx",
#     filetypes=[("Excel files", "*.xlsx")]
# )
# if unmatched_file:
#     unmatched.to_excel(unmatched_file, index=False)
#     print(f"Done! Unmatched data saved to {unmatched_file}")


# python -m PyInstaller --onefile --noconsole sort.py


# import pandas as pd
# from tkinter import Tk, filedialog, StringVar, OptionMenu, Button, Toplevel

# root = Tk()
# root.withdraw()  


# def load_file(path):
#     return pd.read_csv(path) if path.endswith(".csv") else pd.read_excel(path)

# file1 = filedialog.askopenfilename(title="Select first file", filetypes=[("Excel or CSV", "*.xlsx *.csv")])
# file2 = filedialog.askopenfilename(title="Select second file", filetypes=[("Excel or CSV", "*.xlsx *.csv")])

# df1, df2 = load_file(file1), load_file(file2)

# win = Toplevel(root)
# win.title("Select Columns")
# var1, var2 = StringVar(value=df1.columns[0]), StringVar(value=df2.columns[0])
# OptionMenu(win, var1, *df1.columns).pack(padx=10, pady=5)
# OptionMenu(win, var2, *df2.columns).pack(padx=10, pady=5)
# Button(win, text="OK", command=win.destroy).pack(pady=10)
# root.wait_window(win)

# col1, col2 = var1.get(), var2.get()

# unmatched = pd.merge(df1, df2, left_on=col1, right_on=col2, how="outer", indicator=True)
# unmatched = unmatched[unmatched["_merge"] != "both"]

# save_path = filedialog.asksaveasfilename(title="Save unmatched file", defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
# if save_path:
#     unmatched.to_excel(save_path, index=False)
#     print("Done! File saved:", save_path)




# import pandas as pd
# from tkinter import Tk, filedialog, StringVar, OptionMenu, Button, Toplevel

# root = Tk(); root.withdraw()

# def load_file(path):
#     return pd.read_csv(path) if path.endswith(".csv") else pd.read_excel(path)

# f1 = filedialog.askopenfilename(title="Select first file", filetypes=[("Excel or CSV", "*.xlsx *.csv")])
# f2 = filedialog.askopenfilename(title="Select second file", filetypes=[("Excel or CSV", "*.xlsx *.csv")])
# df1, df2 = load_file(f1), load_file(f2)

# win = Toplevel(root); win.title("Select Columns")
# c1, c2 = StringVar(value=df1.columns[0]), StringVar(value=df2.columns[0])
# OptionMenu(win, c1, *df1.columns).pack(padx=10, pady=5)
# OptionMenu(win, c2, *df2.columns).pack(padx=10, pady=5)
# Button(win, text="OK", command=win.destroy).pack(pady=10)
# root.wait_window(win)

# unmatched = pd.merge(df1, df2, left_on=c1.get(), right_on=c2.get(), how="outer", indicator=True)
# unmatched = unmatched[unmatched["_merge"] != "both"]

# cols1 = df1.columns.tolist()
# cols2 = [c for c in unmatched.columns if c not in cols1 + ["_merge"]]
# unmatched = unmatched.reindex(columns=cols1 + ["", ""] + cols2)

# path = filedialog.asksaveasfilename(title="Save unmatched file", defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
# if path:
#     unmatched.to_excel(path, index=False)
#     print("Saved:", path)



import pandas as pd
from tkinter import Tk, filedialog, StringVar, OptionMenu, Button, Toplevel, Label
from openpyxl import load_workbook

root = Tk(); root.withdraw()

def load_file(path): 
    return pd.read_csv(path, dtype=str) if path.endswith(".csv") else pd.read_excel(path, dtype=str)

f1 = filedialog.askopenfilename(title="Select first file", filetypes=[("Excel or CSV", "*.xlsx *.csv")])
f2 = filedialog.askopenfilename(title="Select second file", filetypes=[("Excel or CSV", "*.xlsx *.csv")])
df1, df2 = load_file(f1), load_file(f2)

win = Toplevel(root); win.title("Select Columns")
w,h,sw,sh=400,200,win.winfo_screenwidth(),win.winfo_screenheight()
win.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}"); win.resizable(False,False)
c1,c2 = StringVar(value=df1.columns[0]), StringVar(value=df2.columns[0])
Label(win,text="First file column").pack(); OptionMenu(win,c1,*df1.columns).pack()
Label(win,text="Second file column").pack(); OptionMenu(win,c2,*df2.columns).pack()
Button(win,text="OK",command=win.destroy).pack(pady=10); root.wait_window(win)

unmatched = pd.merge(df1,df2,left_on=c1.get(),right_on=c2.get(),how="outer",indicator=True)
unmatched = unmatched[unmatched["_merge"]!="both"].reindex(columns=df1.columns.tolist()+["",""]+[c for c in unmatched.columns if c not in df1.columns and c!="_merge"])
unmatched = unmatched.astype(str).fillna("").replace("nan","")

path = filedialog.asksaveasfilename(title="Save unmatched file",defaultextension=".xlsx",filetypes=[("Excel","*.xlsx")])
if path:
    unmatched.to_excel(path,index=False,engine="openpyxl")
    wb=load_workbook(path); ws=wb.active
    [[setattr(c,"number_format","@") for c in r] for r in ws.iter_rows()]
    wb.save(path)
    print("Saved:",path)
